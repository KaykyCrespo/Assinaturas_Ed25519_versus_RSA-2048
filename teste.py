#!/usr/bin/env python3
"""
Análise do Impacto do Tamanho de Blocos de Dados no Desempenho 
da Verificação de Assinaturas Ed25519 versus RSA-2048

Autor: [Seu Nome]
Data: Agosto 2024
"""

import os
import time
import json
import statistics
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.asymmetric import rsa, ed25519, padding
from cryptography.exceptions import InvalidSignature
from scipy import stats
from sklearn.linear_model import LinearRegression
import platform
import warnings
warnings.filterwarnings('ignore')

class VerificationBenchmark:
    """
    Classe para benchmark de verificação de assinaturas Ed25519 vs RSA-2048
    com diferentes tamanhos de blocos de dados.
    """
    
    def __init__(self):
        """Inicializa o benchmark com configurações padrão."""
        # Tamanhos de blocos para teste (em bytes)
        self.block_sizes = {
            '1KB': 1024,
            '10KB': 10240,
            '100KB': 102400,
            '1MB': 1048576,
            '10MB': 10485760,
            '50MB': 52428800,
            '100MB': 104857600
        }
        
        # Número de repetições para cada teste
        self.iterations = 100  # Ajustável conforme necessário
        
        # Resultados
        self.results = []
        
        # Preparar chaves
        print("Inicializando ambiente de teste...")
        self._prepare_keys()
        
    def _prepare_keys(self):
        """Gera as chaves para Ed25519 e RSA-2048."""
        print("Gerando chaves criptográficas...")
        
        # Ed25519
        self.ed25519_private = ed25519.Ed25519PrivateKey.generate()
        self.ed25519_public = self.ed25519_private.public_key()
        
        # RSA-2048
        self.rsa_private = rsa.generate_private_key(
            public_exponent=65537,
            key_size=2048
        )
        self.rsa_public = self.rsa_private.public_key()
        
        print("✓ Chaves geradas com sucesso")
    
    def benchmark_verification(self, algorithm, data_size_bytes):
        """
        Realiza benchmark de verificação para um algoritmo e tamanho específicos.
        
        Args:
            algorithm: 'Ed25519' ou 'RSA-2048'
            data_size_bytes: Tamanho dos dados em bytes
            
        Returns:
            Dict com estatísticas do benchmark
        """
        # Gerar dados aleatórios
        data = os.urandom(data_size_bytes)
        
        # Criar assinatura
        if algorithm == 'Ed25519':
            signature = self.ed25519_private.sign(data)
            public_key = self.ed25519_public
        else:  # RSA-2048
            signature = self.rsa_private.sign(
                data,
                padding.PSS(
                    mgf=padding.MGF1(hashes.SHA256()),
                    salt_length=padding.PSS.MAX_LENGTH
                ),
                hashes.SHA256()
            )
            public_key = self.rsa_public
        
        # Realizar múltiplas verificações e medir tempo
        times = []
        
        # Aquecimento (descartado)
        for _ in range(10):
            if algorithm == 'Ed25519':
                public_key.verify(signature, data)
            else:
                public_key.verify(
                    signature, data,
                    padding.PSS(
                        mgf=padding.MGF1(hashes.SHA256()),
                        salt_length=padding.PSS.MAX_LENGTH
                    ),
                    hashes.SHA256()
                )
        
        # Medições reais
        for _ in range(self.iterations):
            start = time.perf_counter_ns()
            
            if algorithm == 'Ed25519':
                public_key.verify(signature, data)
            else:  # RSA-2048
                public_key.verify(
                    signature, data,
                    padding.PSS(
                        mgf=padding.MGF1(hashes.SHA256()),
                        salt_length=padding.PSS.MAX_LENGTH
                    ),
                    hashes.SHA256()
                )
            
            end = time.perf_counter_ns()
            times.append((end - start) / 1_000_000)  # Converter para ms
        
        # Calcular estatísticas
        return {
            'algorithm': algorithm,
            'data_size_bytes': data_size_bytes,
            'data_size_mb': data_size_bytes / (1024 * 1024),
            'mean_ms': statistics.mean(times),
            'median_ms': statistics.median(times),
            'stdev_ms': statistics.stdev(times),
            'min_ms': min(times),
            'max_ms': max(times),
            'p95_ms': np.percentile(times, 95),
            'p99_ms': np.percentile(times, 99),
            'throughput_mbps': (data_size_bytes / (1024 * 1024)) / (statistics.mean(times) / 1000),
            'all_times': times
        }
    
    def run_complete_benchmark(self):
        """Executa o benchmark completo para todos os tamanhos e algoritmos."""
        print("\n" + "="*60)
        print("INICIANDO BENCHMARK DE VERIFICAÇÃO")
        print("="*60)
        
        total_tests = len(self.block_sizes) * 2  # 2 algoritmos
        current_test = 0
        
        for size_label, size_bytes in self.block_sizes.items():
            print(f"\n→ Testando bloco de {size_label}...")
            
            # Ed25519
            current_test += 1
            print(f"  [{current_test}/{total_tests}] Ed25519... ", end='', flush=True)
            result_ed = self.benchmark_verification('Ed25519', size_bytes)
            result_ed['size_label'] = size_label
            self.results.append(result_ed)
            print(f"✓ {result_ed['mean_ms']:.3f}ms")
            
            # RSA-2048
            current_test += 1
            print(f"  [{current_test}/{total_tests}] RSA-2048... ", end='', flush=True)
            result_rsa = self.benchmark_verification('RSA-2048', size_bytes)
            result_rsa['size_label'] = size_label
            self.results.append(result_rsa)
            print(f"✓ {result_rsa['mean_ms']:.3f}ms")
            
            # Mostrar comparação imediata
            ratio = result_rsa['mean_ms'] / result_ed['mean_ms']
            print(f"  → RSA/Ed25519 ratio: {ratio:.2f}x")
        
        print("\n" + "="*60)
        print("BENCHMARK CONCLUÍDO")
        print("="*60)
    
    def analyze_results(self):
        """Analisa os resultados e identifica pontos importantes."""
        df = pd.DataFrame(self.results)
        
        print("\n" + "="*60)
        print("ANÁLISE DOS RESULTADOS")
        print("="*60)
        
        # Tabela comparativa
        pivot_table = df.pivot(index='size_label', columns='algorithm', values='mean_ms')
        pivot_table['Ratio (RSA/Ed)'] = pivot_table['RSA-2048'] / pivot_table['Ed25519']
        
        print("\n📊 Tempo Médio de Verificação (ms):")
        print("-" * 50)
        print(pivot_table.to_string(float_format='%.3f'))
        
        # Análise de throughput
        pivot_throughput = df.pivot(index='size_label', columns='algorithm', values='throughput_mbps')
        print("\n📈 Throughput (MB/s):")
        print("-" * 50)
        print(pivot_throughput.to_string(float_format='%.2f'))
        
        # Encontrar ponto de cruzamento (se existir)
        self._find_crossover_point(df)
        
        # Modelagem matemática
        self._create_predictive_model(df)
        
        return df
    
    def _find_crossover_point(self, df):
        """Identifica o ponto onde as curvas de desempenho se aproximam."""
        print("\n🎯 Análise de Convergência:")
        print("-" * 50)
        
        ed_data = df[df['algorithm'] == 'Ed25519'].sort_values('data_size_mb')
        rsa_data = df[df['algorithm'] == 'RSA-2048'].sort_values('data_size_mb')
        
        ratios = []
        for size in ed_data['data_size_mb'].values:
            ed_time = ed_data[ed_data['data_size_mb'] == size]['mean_ms'].values[0]
            rsa_time = rsa_data[rsa_data['data_size_mb'] == size]['mean_ms'].values[0]
            ratio = rsa_time / ed_time
            ratios.append((size, ratio))
            
        # Encontrar onde a razão é mínima
        min_ratio = min(ratios, key=lambda x: x[1])
        print(f"Menor diferença relativa: {min_ratio[1]:.2f}x em {min_ratio[0]:.1f}MB")
        
        # Verificar tendência
        if ratios[-1][1] < ratios[0][1]:
            print("Tendência: Convergência com aumento do tamanho dos dados")
        else:
            print("Tendência: Divergência com aumento do tamanho dos dados")
    
    def _create_predictive_model(self, df):
        """Cria modelos preditivos para cada algoritmo."""
        print("\n📐 Modelos Preditivos (T = a + b*S):")
        print("-" * 50)
        
        for algo in ['Ed25519', 'RSA-2048']:
            algo_data = df[df['algorithm'] == algo]
            X = algo_data['data_size_mb'].values.reshape(-1, 1)
            y = algo_data['mean_ms'].values
            
            model = LinearRegression()
            model.fit(X, y)
            
            r2 = model.score(X, y)
            
            print(f"\n{algo}:")
            print(f"  T = {model.intercept_:.3f} + {model.coef_[0]:.3f} × S")
            print(f"  R² = {r2:.4f}")
            
            # Predições para tamanhos intermediários
            test_sizes = [5, 25, 75]  # MB
            print(f"  Predições:")
            for size in test_sizes:
                pred = model.predict([[size]])[0]
                print(f"    {size}MB: {pred:.2f}ms")
    
    def generate_plots(self):
        """Gera gráficos para visualização dos resultados."""
        df = pd.DataFrame(self.results)
        
        # Configurar estilo
        plt.style.use('seaborn-v0_8-darkgrid')
        fig, axes = plt.subplots(2, 2, figsize=(14, 10))
        
        # Gráfico 1: Tempo de execução vs Tamanho
        ax1 = axes[0, 0]
        for algo in ['Ed25519', 'RSA-2048']:
            algo_data = df[df['algorithm'] == algo].sort_values('data_size_mb')
            ax1.plot(algo_data['data_size_mb'], algo_data['mean_ms'], 
                    marker='o', linewidth=2, markersize=8, label=algo)
        ax1.set_xlabel('Tamanho do Bloco (MB)')
        ax1.set_ylabel('Tempo de Verificação (ms)')
        ax1.set_title('Tempo de Verificação vs Tamanho do Bloco')
        ax1.legend()
        ax1.grid(True, alpha=0.3)
        ax1.set_xscale('log')
        ax1.set_yscale('log')
        
        # Gráfico 2: Throughput
        ax2 = axes[0, 1]
        for algo in ['Ed25519', 'RSA-2048']:
            algo_data = df[df['algorithm'] == algo].sort_values('data_size_mb')
            ax2.plot(algo_data['data_size_mb'], algo_data['throughput_mbps'], 
                    marker='s', linewidth=2, markersize=8, label=algo)
        ax2.set_xlabel('Tamanho do Bloco (MB)')
        ax2.set_ylabel('Throughput (MB/s)')
        ax2.set_title('Throughput vs Tamanho do Bloco')
        ax2.legend()
        ax2.grid(True, alpha=0.3)
        ax2.set_xscale('log')
        
        # Gráfico 3: Razão RSA/Ed25519
        ax3 = axes[1, 0]
        ed_data = df[df['algorithm'] == 'Ed25519'].sort_values('data_size_mb')
        rsa_data = df[df['algorithm'] == 'RSA-2048'].sort_values('data_size_mb')
        
        ratios = []
        sizes = []
        for size in ed_data['data_size_mb'].values:
            ed_time = ed_data[ed_data['data_size_mb'] == size]['mean_ms'].values[0]
            rsa_time = rsa_data[rsa_data['data_size_mb'] == size]['mean_ms'].values[0]
            ratios.append(rsa_time / ed_time)
            sizes.append(size)
        
        ax3.plot(sizes, ratios, marker='D', linewidth=2, markersize=8, color='red')
        ax3.axhline(y=1, color='gray', linestyle='--', alpha=0.5)
        ax3.set_xlabel('Tamanho do Bloco (MB)')
        ax3.set_ylabel('Razão RSA/Ed25519')
        ax3.set_title('Vantagem Relativa do Ed25519')
        ax3.grid(True, alpha=0.3)
        ax3.set_xscale('log')
        
        # Gráfico 4: Box plot comparativo
        ax4 = axes[1, 1]
        data_for_box = []
        labels_for_box = []
        
        for size_label in ['1KB', '100KB', '10MB', '100MB']:
            for algo in ['Ed25519', 'RSA-2048']:
                result = next((r for r in self.results 
                             if r['size_label'] == size_label and r['algorithm'] == algo), None)
                if result:
                    data_for_box.append(result['all_times'])
                    labels_for_box.append(f"{algo}\n{size_label}")
        
        bp = ax4.boxplot(data_for_box, labels=labels_for_box, patch_artist=True)
        colors = ['lightblue', 'lightcoral'] * 4
        for patch, color in zip(bp['boxes'], colors):
            patch.set_facecolor(color)
        
        ax4.set_ylabel('Tempo de Verificação (ms)')
        ax4.set_title('Distribuição dos Tempos de Verificação')
        ax4.grid(True, alpha=0.3, axis='y')
        
        plt.tight_layout()
        
        # Salvar figura
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"benchmark_results_{timestamp}.png"
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        print(f"\n📊 Gráficos salvos em: {filename}")
        
        plt.show()
    
    def export_results(self, format='all'):
        """
        Exporta os resultados para diferentes formatos.
        
        Args:
            format: 'csv', 'json', 'excel' ou 'all'
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        df = pd.DataFrame(self.results)
        
        # Remover a coluna 'all_times' para exportação
        df_export = df.drop('all_times', axis=1)
        
        if format in ['csv', 'all']:
            filename = f"results_{timestamp}.csv"
            df_export.to_csv(filename, index=False)
            print(f"✓ Resultados exportados para {filename}")
        
        if format in ['json', 'all']:
            filename = f"results_{timestamp}.json"
            df_export.to_json(filename, orient='records', indent=2)
            print(f"✓ Resultados exportados para {filename}")
        
        if format in ['excel', 'all']:
            # Importar openpyxl para formatação avançada
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.formatting.rule import ColorScaleRule
            import platform
            
            filename = f"Resultados_Ed25519_vs_RSA2048_{timestamp}.xlsx"
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # === SHEET 1: RESULTADOS PRINCIPAIS ===
                df_export.to_excel(writer, sheet_name='Resultados Benchmark', index=False, startrow=8)
                ws = writer.sheets['Resultados Benchmark']
                
                # Estilos
                header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                title_font = Font(name='Calibri', size=16, bold=True, color='366092')
                subtitle_font = Font(name='Calibri', size=11, italic=True)
                
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Título e informações do sistema
                ws['A1'] = "RELATÓRIO DE BENCHMARK CRIPTOGRÁFICO"
                ws['A1'].font = title_font
                ws.merge_cells('A1:K1')
                
                ws['A3'] = f"Data de Execução: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
                ws['A4'] = f"Sistema: {platform.system()} {platform.version()}"
                ws['A5'] = f"Algoritmos Testados: Ed25519 vs RSA-2048"
                ws['A6'] = f"Tamanhos de Bloco: {', '.join(self.block_sizes.keys())}"
                ws['A7'] = f"Iterações por Teste: {self.iterations}"
                
                for row in range(3, 8):
                    ws[f'A{row}'].font = subtitle_font
                
                # Formatar cabeçalhos
                for col in range(1, 12):
                    cell = ws.cell(row=9, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # Formatar dados
                light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                data_font = Font(name='Calibri', size=11)
                data_alignment = Alignment(horizontal='center', vertical='center')
                
                for row in range(10, 10 + len(df_export)):
                    for col in range(1, 12):
                        cell = ws.cell(row=row, column=col)
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = thin_border
                        
                        # Formatação numérica
                        if col in [4, 5, 6, 7, 8, 9, 10, 11]:  # Colunas numéricas
                            if col in [4, 11]:  # MB e throughput
                                cell.number_format = '0.00'
                            else:  # Tempos em ms
                                cell.number_format = '0.0000'
                        
                        # Cores alternadas
                        if (row - 10) % 2 == 1:
                            cell.fill = light_fill
                
                # Ajustar larguras das colunas
                column_widths = {
                    'A': 12,  # algorithm
                    'B': 12,  # size_label
                    'C': 18,  # data_size_bytes
                    'D': 16,  # data_size_mb
                    'E': 16,  # mean_ms
                    'F': 16,  # median_ms
                    'G': 14,  # stdev_ms
                    'H': 12,  # min_ms
                    'I': 12,  # max_ms
                    'J': 12,  # p95_ms
                    'K': 12,  # p99_ms
                    'L': 18   # throughput_mbps
                }
                
                for col, width in column_widths.items():
                    if col <= 'L':
                        ws.column_dimensions[col].width = width
                
                # Formatação condicional para tempo de execução
                if len(df_export) > 1:
                    ws.conditional_formatting.add(
                        f'E10:E{9+len(df_export)}',
                        ColorScaleRule(
                            start_type='min', start_color='63BE7B',
                            mid_type='percentile', mid_value=50, mid_color='FFEB84',
                            end_type='max', end_color='F8696B'
                        )
                    )
                
                # === SHEET 2: TABELA COMPARATIVA ===
                pivot = df.pivot(index='size_label', columns='algorithm', values='mean_ms')
                pivot['Ratio (RSA/Ed)'] = pivot['RSA-2048'] / pivot['Ed25519']
                pivot['Speedup Ed25519'] = pivot['RSA-2048'] / pivot['Ed25519']
                
                pivot.to_excel(writer, sheet_name='Análise Comparativa', startrow=8)
                ws2 = writer.sheets['Análise Comparativa']
                
                # Título
                ws2['A1'] = "ANÁLISE COMPARATIVA DE DESEMPENHO"
                ws2['A1'].font = title_font
                ws2.merge_cells('A1:E1')
                
                ws2['A3'] = "Tempo Médio de Verificação (ms)"
                ws2['A4'] = f"Base: {self.iterations} iterações por teste"
                ws2['A3'].font = Font(name='Calibri', size=12, bold=True)
                ws2['A4'].font = subtitle_font
                
                # Formatar cabeçalhos
                for col in range(1, 6):
                    cell = ws2.cell(row=9, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # Formatar dados com cores para cada algoritmo
                ed_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
                rsa_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
                ratio_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
                
                for row in range(10, 10 + len(pivot)):
                    for col in range(1, 6):
                        cell = ws2.cell(row=row, column=col)
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = thin_border
                        
                        if col == 2:  # Ed25519
                            cell.fill = ed_fill
                            cell.number_format = '0.0000'
                        elif col == 3:  # RSA-2048
                            cell.fill = rsa_fill
                            cell.number_format = '0.0000'
                        elif col in [4, 5]:  # Ratios
                            cell.fill = ratio_fill
                            cell.number_format = '0.00'
                
                # Ajustar larguras
                ws2.column_dimensions['A'].width = 15
                ws2.column_dimensions['B'].width = 15
                ws2.column_dimensions['C'].width = 15
                ws2.column_dimensions['D'].width = 18
                ws2.column_dimensions['E'].width = 18
                
                # === SHEET 3: RESUMO ESTATÍSTICO ===
                stats_data = []
                for algo in ['Ed25519', 'RSA-2048']:
                    algo_df = df[df['algorithm'] == algo]
                    stats_data.append({
                        'Algoritmo': algo,
                        'Tempo Médio (ms)': algo_df['mean_ms'].mean(),
                        'Tempo Mediano (ms)': algo_df['median_ms'].median(),
                        'Tempo Mínimo (ms)': algo_df['mean_ms'].min(),
                        'Tempo Máximo (ms)': algo_df['mean_ms'].max(),
                        'Desvio Padrão Médio': algo_df['stdev_ms'].mean(),
                        'Throughput Médio (MB/s)': algo_df['throughput_mbps'].mean(),
                        'Throughput Máximo (MB/s)': algo_df['throughput_mbps'].max()
                    })
                
                stats_summary = pd.DataFrame(stats_data)
                stats_summary.to_excel(writer, sheet_name='Resumo Estatístico', index=False, startrow=8)
                ws3 = writer.sheets['Resumo Estatístico']
                
                # Título
                ws3['A1'] = "RESUMO ESTATÍSTICO POR ALGORITMO"
                ws3['A1'].font = title_font
                ws3.merge_cells('A1:H1')
                
                ws3['A3'] = "Métricas de Desempenho Consolidadas"
                ws3['A4'] = f"Dados de {len(self.block_sizes)} tamanhos diferentes de blocos"
                ws3['A3'].font = Font(name='Calibri', size=12, bold=True)
                ws3['A4'].font = subtitle_font
                
                # Formatar cabeçalhos e dados
                for col in range(1, 9):
                    # Cabeçalho
                    cell = ws3.cell(row=9, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                    
                    # Dados
                    for row in range(10, 12):
                        cell = ws3.cell(row=row, column=col)
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = thin_border
                        
                        if col > 1:  # Colunas numéricas
                            cell.number_format = '0.0000'
                        
                        # Colorir linha do Ed25519
                        if row == 10:
                            cell.fill = ed_fill
                        else:
                            cell.fill = rsa_fill
                
                # Ajustar larguras
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    ws3.column_dimensions[col].width = 20
                
                # Congelar painéis
                ws.freeze_panes = 'A10'
                ws2.freeze_panes = 'A10'
                ws3.freeze_panes = 'A10'
            
            print(f"✓ Resultados exportados para {filename} (Excel formatado)")
    
    def generate_report(self):
        """Gera um relatório em texto dos resultados."""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        df = pd.DataFrame(self.results)
        
        report = []
        report.append("="*70)
        report.append("RELATÓRIO DE BENCHMARK DE VERIFICAÇÃO DE ASSINATURAS")
        report.append("="*70)
        report.append(f"Data de Execução: {timestamp}")
        report.append(f"Algoritmos Testados: Ed25519, RSA-2048")
        report.append(f"Tamanhos de Bloco: {', '.join(self.block_sizes.keys())}")
        report.append(f"Iterações por Teste: {self.iterations}")
        report.append("")
        
        # Resultados principais
        report.append("RESULTADOS PRINCIPAIS")
        report.append("-"*70)
        
        pivot = df.pivot(index='size_label', columns='algorithm', values='mean_ms')
        pivot['Ratio'] = pivot['RSA-2048'] / pivot['Ed25519']
        
        report.append("\nTempo Médio de Verificação (ms):")
        report.append(pivot.to_string(float_format='%.3f'))
        
        # Análise de desempenho
        report.append("\n\nANÁLISE DE DESEMPENHO")
        report.append("-"*70)
        
        # Melhor e pior caso para cada algoritmo
        for algo in ['Ed25519', 'RSA-2048']:
            algo_data = df[df['algorithm'] == algo]
            best = algo_data.loc[algo_data['mean_ms'].idxmin()]
            worst = algo_data.loc[algo_data['mean_ms'].idxmax()]
            
            report.append(f"\n{algo}:")
            report.append(f"  Melhor caso: {best['size_label']} - {best['mean_ms']:.3f}ms")
            report.append(f"  Pior caso: {worst['size_label']} - {worst['mean_ms']:.3f}ms")
            report.append(f"  Throughput médio: {algo_data['throughput_mbps'].mean():.2f} MB/s")
        
        # Recomendações
        report.append("\n\nRECOMENDAÇÕES")
        report.append("-"*70)
        
        # Análise por faixa de tamanho
        small_data = df[df['data_size_mb'] < 1]
        medium_data = df[(df['data_size_mb'] >= 1) & (df['data_size_mb'] <= 10)]
        large_data = df[df['data_size_mb'] > 10]
        
        for data, label in [(small_data, "Dados Pequenos (<1MB)"),
                            (medium_data, "Dados Médios (1-10MB)"),
                            (large_data, "Dados Grandes (>10MB)")]:
            if not data.empty:
                ed_mean = data[data['algorithm'] == 'Ed25519']['mean_ms'].mean()
                rsa_mean = data[data['algorithm'] == 'RSA-2048']['mean_ms'].mean()
                ratio = rsa_mean / ed_mean
                
                report.append(f"\n{label}:")
                report.append(f"  Vantagem do Ed25519: {ratio:.2f}x")
                
                if ratio > 5:
                    report.append("  Recomendação: Ed25519 FORTEMENTE recomendado")
                elif ratio > 2:
                    report.append("  Recomendação: Ed25519 recomendado")
                else:
                    report.append("  Recomendação: Escolha baseada em outros fatores")
        
        # Salvar relatório
        filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        with open(filename, 'w', encoding='utf-8') as f:
            f.write('\n'.join(report))
        
        print(f"\n📄 Relatório salvo em: {filename}")
        
        # Também imprimir na tela
        print('\n'.join(report))


def main():
    """Função principal para executar o benchmark."""
    print("\n" + "="*70)
    print(" BENCHMARK DE VERIFICAÇÃO DE ASSINATURAS DIGITAIS")
    print(" Ed25519 vs RSA-2048: Análise por Tamanho de Bloco")
    print("="*70)
    
    # Criar e executar benchmark
    benchmark = VerificationBenchmark()
    
    # Executar testes
    start_time = time.time()
    benchmark.run_complete_benchmark()
    elapsed = time.time() - start_time
    
    print(f"\n⏱️  Tempo total de execução: {elapsed:.2f} segundos")
    
    # Analisar resultados
    df_results = benchmark.analyze_results()
    
    # Gerar visualizações
    benchmark.generate_plots()
    
    # Exportar resultados
    benchmark.export_results('all')
    
    # Gerar relatório
    benchmark.generate_report()
    
    print("\n✅ Benchmark concluído com sucesso!")
    print("📁 Verifique os arquivos gerados no diretório atual.")


if __name__ == "__main__":
    main()